from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


FILES = {
    "desktop_ref": Path(r"C:\Users\Leo.Li\Desktop\c4c_ticket_table_z007_z010_with_invoice_layout_checked.xlsx"),
    "web_source": Path(r"C:\Users\Leo.Li\Documents\ChatGPT\Service centre dashboard\c4c_ticket_table_z007_z010_checked_hana_final.xlsx"),
    "latest_export": Path(
        r"C:\Users\Leo.Li\Documents\GitHub\Service-center-dashboard\outputs\01a01c4f-25b7-7e91-8ffd-41a61b3fc906\c4c_ticket_table_z007_z010_with_invoice_layout_checked_latest_like_reference.xlsx"
    ),
}


def future_summary(path: Path) -> dict:
    frame = pd.read_excel(path, sheet_name="Tickets", dtype=str).fillna("")
    created = pd.to_datetime(frame["CreatedOn"], errors="coerce", dayfirst=False)
    cutoff = pd.Timestamp("2026-08-20")
    future = frame[created > cutoff].copy()
    return {
        "rows": int(len(frame)),
        "createdMin": str(created.min()),
        "createdMax": str(created.max()),
        "futureAfter2026_08_20": int(len(future)),
        "futureByMonth": created[created > cutoff].dt.to_period("M").astype(str).value_counts().to_dict(),
        "futureStatus": future["TicketStatusText"].value_counts().to_dict(),
        "futureType": future["TicketTypeText"].value_counts().to_dict(),
        "futureSample": future[
            ["TicketID", "TicketTypeText", "DealerName", "TicketStatusText", "CreatedOn"]
        ].head(15).to_dict(orient="records"),
    }


def openpyxl_samples(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Tickets"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {header: idx for idx, header in enumerate(headers)}
    targets = {"39010", "39066", "39077"}
    rows = []
    for row in sheet.iter_rows(min_row=2):
        ticket_id = str(row[index["TicketID"]].value).strip()
        if ticket_id in targets:
            created_cell = row[index["CreatedOn"]]
            rows.append(
                {
                    "TicketID": ticket_id,
                    "CreatedOnValue": str(created_cell.value),
                    "CreatedOnType": type(created_cell.value).__name__,
                    "CreatedOnNumberFormat": created_cell.number_format,
                    "TicketStatusText": row[index["TicketStatusText"]].value,
                }
            )
    return rows


def main() -> None:
    output = {name: future_summary(path) for name, path in FILES.items()}
    output["webSourceOpenpyxlSamples"] = openpyxl_samples(FILES["web_source"])
    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
