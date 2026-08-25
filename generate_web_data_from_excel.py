from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
SOURCE_WORKBOOK = BASE_DIR / "c4c_ticket_table_z007_z010_checked_hana_final.xlsx"
OUTPUT_JS = BASE_DIR / "dashboard-data.js"
ABNORMAL_WORKBOOK = BASE_DIR / "abnormal_tickets.xlsx"

DEALER_LABELS = {
    "Regent RV - Perth": "Perth",
    "Regent RV - Traralgon": "Traralgon",
    "Snowy River Launceston": "Launceston",
    "Snowy River Geelong": "Geelong",
    "Regent RV - Frankston": "Frankston",
}

DEALER_ORDER = ["Perth", "Traralgon", "Launceston", "Geelong", "Frankston", "Other"]
DEALER_COLORS = {
    "Perth": "#1f6feb",
    "Traralgon": "#17a6ad",
    "Launceston": "#f58b1f",
    "Geelong": "#7c3aed",
    "Frankston": "#22a447",
    "Other": "#64748b",
}

STATUS_COLORS = [
    "#1f6feb",
    "#17a6ad",
    "#f58b1f",
    "#7c3aed",
    "#22a447",
    "#9a3412",
    "#64748b",
    "#ef4444",
]

TICKET_TYPE_FILTERS = ["All Ticket Types", "Repair ticket", "PDI ticket"]

CANCELLED_STATUS_TEXT = {"cancel", "cancel invoice"}
CREATE_INVOICE_STATUS = "create invoice"
INTERNAL_PARTY_NAME = "REGENT RV PTY LTD"
ABNORMAL_FILL = PatternFill(fill_type="solid", fgColor="FDE2E2")
AGING_BUCKETS = [
    ("0-7 days", 0, 7),
    ("8-30 days", 8, 30),
    ("31-60 days", 31, 60),
    ("60+ days", 61, None),
]
REPAIR_WORKFLOW_STAGES = [
    "Awaiting Quote Approval",
    "Approved, Awaiting Repair",
    "Repair In Progress",
    "Repair Complete, Awaiting Time Claim",
    "Time Claimed, Awaiting Invoice",
    "Completed / Invoiced",
    "Invoice Cancelled, SO Still Open",
    "Cancelled",
]
PDI_WORKFLOW_STAGES = [
    "Awaiting PDI Start",
    "PDI Complete, Awaiting Time Claim",
    "PDI Completed",
    "Cancelled",
]
STAGE_COLOR_OVERRIDES = {
    "Awaiting Quote Approval": "#1f6feb",
    "Approved, Awaiting Repair": "#17a6ad",
    "Repair In Progress": "#f58b1f",
    "Repair Complete, Awaiting Time Claim": "#7c3aed",
    "Time Claimed, Awaiting Invoice": "#22a447",
    "Completed / Invoiced": "#0f766e",
    "Invoice Cancelled, SO Still Open": "#9a3412",
    "Cancelled": "#64748b",
    "Awaiting PDI Start": "#1f6feb",
    "PDI Complete, Awaiting Time Claim": "#17a6ad",
    "PDI Completed": "#0f766e",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def parse_amount(series: pd.Series) -> pd.Series:
    cleaned = (
        series.fillna("")
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)


def parse_hours(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series.fillna("").astype(str).str.strip(), errors="coerce").fillna(0.0)


def money(value: float) -> str:
    if abs(value) >= 1_000_000:
        return f"${value / 1_000_000:.2f}M"
    if abs(value) >= 1_000:
        return f"${value / 1_000:.1f}K"
    return f"${value:.0f}"


def party_name_normalized(row: pd.Series) -> str:
    return clean(row.get("Role_1001_InvolvedPartyName", "")).upper()


def is_internal_party(row: pd.Series) -> bool:
    return party_name_normalized(row) == INTERNAL_PARTY_NAME


def is_abnormal_party(row: pd.Series) -> bool:
    party_id = clean(row.get("Role_1001_InvolvedPartyID", "")).upper()
    party_name = party_name_normalized(row)
    if not party_name:
        return False
    return party_name != INTERNAL_PARTY_NAME and party_id.endswith(".ORG")


def invoice_scope(row: pd.Series) -> str:
    if is_internal_party(row):
        return "Internal"
    return "External"


def abnormal_reasons(row: pd.Series) -> list[str]:
    reasons = []
    status = clean(row.get("TicketStatusText", "")).lower()
    has_invoice_number = bool(clean(row.get("ERPInvoiceNumber", "")))
    has_invoice_price = bool(clean(row.get("ERPInvoiceNumberPrice", "")))
    if row.get("AbnormalParty", False):
        reasons.append("Invalid Role_1001 party: dealer/org used as customer")
    if status == CREATE_INVOICE_STATUS and not (has_invoice_number and has_invoice_price):
        reasons.append("Create invoice status but invoice number/price is empty")
    if status != CREATE_INVOICE_STATUS and has_invoice_number:
        reasons.append("Invoice exists but status is not Create invoice")
    return reasons


def is_open_ticket(row: pd.Series) -> bool:
    ticket_type = clean(row.get("TicketType", "")).upper()
    ticket_type_text = clean(row.get("TicketTypeText", "")).lower()
    status = clean(row.get("TicketStatusText", "")).lower()
    is_pdi = ticket_type == "Z010" or ticket_type_text == "pdi"
    is_repair = ticket_type == "Z007" or ticket_type_text == "repair ticket"
    if is_pdi:
        return status not in {"claim time ticket", "cancel"}
    if is_repair:
        return status not in {"cancel", "create invoice"}
    return status not in {"cancel", "cancel invoice", "create invoice", "claim time ticket"}


def ticket_type_bucket(row: pd.Series) -> str:
    ticket_type = clean(row.get("TicketType", "")).upper()
    ticket_type_text = clean(row.get("TicketTypeText", "")).lower()
    if ticket_type == "Z010" or ticket_type_text == "pdi":
        return "PDI ticket"
    if ticket_type == "Z007" or ticket_type_text == "repair ticket":
        return "Repair ticket"
    return "Other"


def workflow_stage(row: pd.Series) -> str:
    status = clean(row.get("TicketStatusText", ""))
    status_key = status.lower()
    ticket_type = ticket_type_bucket(row)
    if ticket_type == "PDI ticket":
        return {
            "open": "Awaiting PDI Start",
            "repair completed": "PDI Complete, Awaiting Time Claim",
            "claim time ticket": "PDI Completed",
            "cancel": "Cancelled",
        }.get(status_key, status or "Blank")
    if ticket_type == "Repair ticket":
        return {
            "open": "Awaiting Quote Approval",
            "quote approved": "Approved, Awaiting Repair",
            "repair in progress": "Repair In Progress",
            "create delivery note": "Repair In Progress",
            "repair completed": "Repair Complete, Awaiting Time Claim",
            "claim time ticket": "Time Claimed, Awaiting Invoice",
            "create invoice": "Completed / Invoiced",
            "cancel invoice": "Invoice Cancelled, SO Still Open",
            "cancel": "Cancelled",
        }.get(status_key, status or "Blank")
    return status or "Blank"


def workflow_stages_for_type(ticket_type_filter: str | None) -> list[str]:
    if ticket_type_filter == "Repair ticket":
        return REPAIR_WORKFLOW_STAGES
    if ticket_type_filter == "PDI ticket":
        return PDI_WORKFLOW_STAGES
    return []


def stage_color(stage: str, index: int) -> str:
    return STAGE_COLOR_OVERRIDES.get(stage, STATUS_COLORS[index % len(STATUS_COLORS)])


def build_aging_counts(rows: pd.DataFrame) -> list[dict[str, Any]]:
    if rows.empty:
        return [{"label": label, "qty": 0} for label, _, _ in AGING_BUCKETS]
    today = pd.Timestamp.now().normalize()
    ages = (today - rows["CreatedDate"].dt.normalize()).dt.days.fillna(0)
    buckets = []
    for label, start, end in AGING_BUCKETS:
        if end is None:
            qty = int((ages >= start).sum())
        else:
            qty = int(((ages >= start) & (ages <= end)).sum())
        buckets.append({"label": label, "qty": qty})
    return buckets


def filter_ticket_type(rows: pd.DataFrame, ticket_type_filter: str) -> pd.DataFrame:
    if ticket_type_filter == "Repair ticket":
        return rows[
            rows["TicketType"].map(lambda value: clean(value).upper()).eq("Z007")
            | rows["TicketTypeText"].map(lambda value: clean(value).lower()).eq("repair ticket")
        ].copy()
    if ticket_type_filter == "PDI ticket":
        return rows[
            rows["TicketType"].map(lambda value: clean(value).upper()).eq("Z010")
            | rows["TicketTypeText"].map(lambda value: clean(value).lower()).eq("pdi")
        ].copy()
    return rows.copy()


def build_dealer_rows(
    created_rows: pd.DataFrame,
    invoice_rows: pd.DataFrame,
    open_rows: pd.DataFrame,
    ticket_type_filter: str | None = None,
) -> list[dict[str, Any]]:
    dealer_rows = []
    for dealer in DEALER_ORDER:
        dealer_tickets = created_rows[created_rows["DealerBucket"].eq(dealer)]
        dealer_invoices = invoice_rows[invoice_rows["DealerBucket"].eq(dealer)]
        dealer_open = open_rows[open_rows["DealerBucket"].eq(dealer)]
        qty = int(len(dealer_tickets))
        quote_amount = float(dealer_tickets["NewTicketQuoteAmount"].sum())
        invoiced_qty = int(len(dealer_invoices))
        invoiced_amount = float(dealer_invoices["InvoiceAmount"].sum())
        open_qty = int(len(dealer_open))
        open_quote_amount = float(dealer_open["NewTicketQuoteAmount"].sum())
        internal_invoices = dealer_invoices[dealer_invoices["InvoiceScope"].eq("Internal")]
        external_invoices = dealer_invoices[dealer_invoices["InvoiceScope"].eq("External")]
        dealer_rows.append(
            {
                "yard": dealer,
                "color": DEALER_COLORS[dealer],
                "newTickets": qty,
                "newQuoteAmount": round(quote_amount, 2),
                "newQuoteAmountLabel": money(quote_amount),
                "newAmount": round(quote_amount, 2),
                "newAmountLabel": money(quote_amount),
                "invoicedTickets": invoiced_qty,
                "invoicedAmount": round(invoiced_amount, 2),
                "invoicedAmountLabel": money(invoiced_amount),
                "openTickets": open_qty,
                "openQuoteAmount": round(open_quote_amount, 2),
                "openQuoteAmountLabel": money(open_quote_amount),
                "internalInvoicedTickets": int(len(internal_invoices)),
                "internalInvoicedAmount": round(float(internal_invoices["InvoiceAmount"].sum()), 2),
                "internalInvoicedAmountLabel": money(float(internal_invoices["InvoiceAmount"].sum())),
                "externalInvoicedTickets": int(len(external_invoices)),
                "externalInvoicedAmount": round(float(external_invoices["InvoiceAmount"].sum()), 2),
                "externalInvoicedAmountLabel": money(float(external_invoices["InvoiceAmount"].sum())),
                "openStatusMix": build_open_status_mix(dealer_open, ticket_type_filter),
            }
        )
    return dealer_rows


def build_invoice_mix(invoice_rows: pd.DataFrame) -> dict[str, Any]:
    total_amount = float(invoice_rows["InvoiceAmount"].sum())
    internal = invoice_rows[invoice_rows["InvoiceScope"].eq("Internal")]
    external = invoice_rows[invoice_rows["InvoiceScope"].eq("External")]
    internal_amount = float(internal["InvoiceAmount"].sum())
    external_amount = float(external["InvoiceAmount"].sum())
    external_percent = round((external_amount / total_amount) * 100) if total_amount else 0
    internal_percent = 100 - external_percent if total_amount else 0
    return {
        "total": money(total_amount),
        "totalAmount": round(total_amount, 2),
        "segments": [
            {
                "name": "External",
                "percent": external_percent,
                "amount": money(external_amount),
                "qty": int(len(external)),
                "color": "#1f6feb",
            },
            {
                "name": "Internal",
                "percent": internal_percent,
                "amount": money(internal_amount),
                "qty": int(len(internal)),
                "color": "#f58b1f",
            },
        ],
    }


def hours_label(value: float) -> str:
    return f"{value:,.1f}h"


def build_labour_summary(rows: pd.DataFrame) -> dict[str, Any]:
    result: dict[str, Any] = {}
    source = rows.copy().reset_index(drop=True)
    source = source.loc[:, ~source.columns.duplicated()].copy()
    if "LabourHours" not in source.columns:
        source["LabourHours"] = 0.0
    if "TicketTypeBucket" not in source.columns:
        source["TicketTypeBucket"] = source.apply(ticket_type_bucket, axis=1) if not source.empty else pd.Series(dtype="object")
    if "WorkerName" not in source.columns:
        source["WorkerName"] = source.get("Role_40_InvolvedPartyName", pd.Series(dtype="object")).map(lambda value: clean(value) or "Unassigned")

    for dealer in DEALER_ORDER:
        dealer_rows = source.loc[source["DealerBucket"].eq(dealer)].copy() if "DealerBucket" in source.columns else source.iloc[0:0].copy()
        repair_rows = dealer_rows.loc[dealer_rows["TicketTypeBucket"].eq("Repair ticket")]
        pdi_rows = dealer_rows.loc[dealer_rows["TicketTypeBucket"].eq("PDI ticket")]
        total_hours = float(dealer_rows["LabourHours"].sum()) if "LabourHours" in dealer_rows.columns else 0.0
        repair_hours = float(repair_rows["LabourHours"].sum()) if "LabourHours" in repair_rows.columns else 0.0
        pdi_hours = float(pdi_rows["LabourHours"].sum()) if "LabourHours" in pdi_rows.columns else 0.0
        worker_rows = []
        if not dealer_rows.empty:
            grouped = dealer_rows.groupby("WorkerName", dropna=False)
            for worker, worker_group in grouped:
                worker_name = clean(worker) or "Unassigned"
                worker_repair = worker_group[worker_group["TicketTypeBucket"].eq("Repair ticket")]
                worker_pdi = worker_group[worker_group["TicketTypeBucket"].eq("PDI ticket")]
                hours = float(worker_group["LabourHours"].sum())
                tickets = int(len(worker_group))
                worker_rows.append(
                    {
                        "worker": worker_name,
                        "tickets": tickets,
                        "hours": round(hours, 2),
                        "hoursLabel": hours_label(hours),
                        "avgHours": round(hours / tickets, 2) if tickets else 0.0,
                        "avgHoursLabel": hours_label(hours / tickets) if tickets else "0.0h",
                        "repairHours": round(float(worker_repair["LabourHours"].sum()), 2),
                        "repairHoursLabel": hours_label(float(worker_repair["LabourHours"].sum())),
                        "pdiHours": round(float(worker_pdi["LabourHours"].sum()), 2),
                        "pdiHoursLabel": hours_label(float(worker_pdi["LabourHours"].sum())),
                    }
                )
        worker_rows.sort(key=lambda item: (-item["hours"], item["worker"]))
        result[dealer] = {
            "totalHours": round(total_hours, 2),
            "totalHoursLabel": hours_label(total_hours),
            "repairHours": round(repair_hours, 2),
            "repairHoursLabel": hours_label(repair_hours),
            "pdiHours": round(pdi_hours, 2),
            "pdiHoursLabel": hours_label(pdi_hours),
            "ticketCount": int(len(dealer_rows)),
            "workerCount": len({row["worker"] for row in worker_rows if row["worker"] != "Unassigned"}),
            "topWorkers": worker_rows[:8],
        }
    return result


def build_open_status_mix(open_rows: pd.DataFrame, ticket_type_filter: str | None = None) -> dict[str, Any]:
    total = int(len(open_rows))
    stage_rows = open_rows.copy()
    if not stage_rows.empty:
        stage_rows["WorkflowStage"] = stage_rows.apply(workflow_stage, axis=1)
        stage_rows["RawTicketStatus"] = stage_rows["TicketStatusText"].map(lambda value: clean(value) or "Blank")
    else:
        stage_rows["WorkflowStage"] = pd.Series(dtype="object")
        stage_rows["RawTicketStatus"] = pd.Series(dtype="object")

    grouped = {}
    if not stage_rows.empty:
        for stage, rows in stage_rows.groupby("WorkflowStage", dropna=False):
            stage_name = clean(stage) or "Blank"
            grouped[stage_name] = {
                "qty": int(len(rows)),
                "quote_amount": float(rows["NewTicketQuoteAmount"].sum()),
                "raw_statuses": sorted(set(rows["RawTicketStatus"])),
                "aging": build_aging_counts(rows),
            }

    stage_order = workflow_stages_for_type(ticket_type_filter)
    if not stage_order:
        stage_order = sorted(grouped, key=lambda stage: (-grouped[stage]["qty"], stage))
    for stage in grouped:
        if stage not in stage_order:
            stage_order.append(stage)

    segments = []
    allocated = 0
    for idx, name in enumerate(stage_order):
        item = grouped.get(name, {})
        qty = int(item.get("qty", 0))
        quote_amount = float(item.get("quote_amount", 0))
        percent = round((qty / total) * 100) if total else 0
        percent_label = "<1%" if total and qty > 0 and (qty / total) * 100 < 1 else f"{percent}%"
        allocated += percent
        segments.append(
            {
                "name": name,
                "share": round((qty / total) * 100, 4) if total else 0,
                "percent": percent,
                "percentLabel": percent_label,
                "qty": qty,
                "quoteAmount": round(quote_amount, 2),
                "quoteAmountLabel": money(quote_amount),
                "amount": f"{qty:,} tickets ({money(quote_amount)})",
                "rawStatuses": item.get("raw_statuses", []),
                "aging": item.get("aging", build_aging_counts(pd.DataFrame(columns=open_rows.columns))),
                "color": stage_color(name, idx),
            }
        )
    if segments:
        segments[-1]["percent"] += 100 - allocated
    return {"total": f"{total:,}", "segments": segments}


def dealer_bucket(dealer_name: str) -> str:
    return DEALER_LABELS.get(clean(dealer_name), "Other")


def month_label(dt: pd.Timestamp) -> str:
    return dt.strftime("%b %Y")


def date_label(value: Any) -> str:
    if pd.isna(value):
        return ""
    return pd.Timestamp(value).strftime("%d/%m/%Y")


def period_label(period: str) -> str:
    if not period:
        return ""
    return pd.Period(period).to_timestamp().strftime("%b %Y")


def numeric_value(value: Any) -> float:
    text = clean(value).replace(",", "").replace("$", "")
    number = pd.to_numeric(pd.Series([text]), errors="coerce").fillna(0.0).iloc[0]
    return float(number)


def build_ticket_details(rows: pd.DataFrame) -> list[dict[str, Any]]:
    details = []
    source = rows.copy().sort_values(["CreatedDate", "TicketID"], ascending=[False, True])
    for _, row in source.iterrows():
        invoice_number = clean(row.get("ERPInvoiceNumber", ""))
        invoice_price = clean(row.get("ERPInvoiceNumberPrice", ""))
        quote_amount = float(row.get("NewTicketQuoteAmount", 0) or 0)
        invoice_amount = float(row.get("InvoiceAmount", 0) or 0)
        labour_hours = float(row.get("LabourHours", 0) or 0)
        claim_hours = numeric_value(row.get("Z1Z8TimeConsumed", ""))
        completed = completed_date(row)
        created_month = clean(row.get("CreatedMonth", ""))
        details.append(
            {
                "serviceOrderId": clean(row.get("TicketID", "")) or "TBC",
                "source": "C4C",
                "period": period_label(created_month),
                "periodKey": created_month,
                "year": clean(row.get("CreatedYear", "")),
                "createdDate": date_label(row.get("CreatedDate")) or "TBC",
                "completedDate": date_label(completed) or "TBC",
                "dealerYard": clean(row.get("DealerBucket", "")) or "Other",
                "dealerName": clean(row.get("DealerName", "")) or "TBC",
                "serviceType": clean(row.get("TicketTypeBucket", "")) or "TBC",
                "ticketTypeCode": clean(row.get("TicketType", "")) or "TBC",
                "status": workflow_stage(row) or "TBC",
                "rawStatus": clean(row.get("TicketStatusText", "")) or "TBC",
                "priority": clean(row.get("TicketSeverity", "")) or "TBC",
                "quoteAmount": round(quote_amount, 2),
                "quoteAmountLabel": money(quote_amount) if quote_amount else "TBC",
                "invoiceNo": invoice_number or "TBC",
                "invoiceAmount": round(invoice_amount, 2),
                "invoiceAmountLabel": money(invoice_amount) if invoice_number and invoice_price else "TBC",
                "billingDate": date_label(row.get("BillingDate")) or "TBC",
                "invoiceScope": clean(row.get("InvoiceScope", "")) or "TBC",
                "labourHours": round(labour_hours, 2),
                "labourHoursLabel": hours_label(labour_hours) if labour_hours else "TBC",
                "claimHours": round(claim_hours, 2),
                "claimHoursLabel": hours_label(claim_hours) if claim_hours else "TBC",
                "actualWorkHours": round(labour_hours, 2),
                "actualWorkHoursLabel": hours_label(labour_hours) if labour_hours else "TBC",
                "invoicePaidHours": "Missing",
                "workerName": clean(row.get("WorkerName", "")) or "TBC",
                "vehicle": clean(row.get("SerialID", "")) or clean(row.get("ChassisNumber", "")) or "TBC",
                "chassisNumber": clean(row.get("ChassisNumber", "")) or "TBC",
                "ticketName": clean(row.get("TicketName", "")) or "TBC",
            }
        )
    return details


def open_rows_at_month_end(open_tickets: pd.DataFrame, period: str) -> pd.DataFrame:
    month_end = pd.Period(period).to_timestamp(how="end")
    return open_tickets[open_tickets["CreatedDate"].le(month_end)].copy()


def open_rows_at_year_end(open_tickets: pd.DataFrame, year: str) -> pd.DataFrame:
    year_end = pd.Period(year, freq="Y").to_timestamp(how="end")
    return open_tickets[open_tickets["CreatedDate"].le(year_end)].copy()


def rows_created_by(rows: pd.DataFrame, end: pd.Timestamp) -> pd.DataFrame:
    return rows[rows["CreatedDate"].le(end)].copy()


def completed_date(row: pd.Series) -> pd.Timestamp | pd.NaT:
    status = clean(row.get("TicketStatusText", "")).lower()
    ticket_type = clean(row.get("TicketType", "")).upper()
    ticket_type_text = clean(row.get("TicketTypeText", "")).lower()
    is_pdi = ticket_type == "Z010" or ticket_type_text == "pdi"
    is_repair = ticket_type == "Z007" or ticket_type_text == "repair ticket"
    if is_repair and status == CREATE_INVOICE_STATUS:
        return row.get("BillingDate") if pd.notna(row.get("BillingDate")) else row.get("ChangeDate")
    if is_pdi and status == "claim time ticket":
        return row.get("ChangeDate") if pd.notna(row.get("ChangeDate")) else row.get("CreatedDate")
    return pd.NaT


def build_status_pipeline(rows: pd.DataFrame, ticket_type_filter: str | None = None) -> list[dict[str, Any]]:
    stage_rows = rows.copy()
    if not stage_rows.empty:
        stage_rows["WorkflowStage"] = stage_rows.apply(workflow_stage, axis=1)
        stage_rows["RawTicketStatus"] = stage_rows["TicketStatusText"].map(lambda value: clean(value) or "Blank")
    else:
        stage_rows["WorkflowStage"] = pd.Series(dtype="object")
        stage_rows["RawTicketStatus"] = pd.Series(dtype="object")
    grouped = {}
    if not stage_rows.empty:
        for stage, stage_group in stage_rows.groupby("WorkflowStage", dropna=False):
            stage_name = clean(stage) or "Blank"
            grouped[stage_name] = {
                "qty": int(len(stage_group)),
                "quote_amount": float(stage_group["NewTicketQuoteAmount"].sum()),
                "raw_statuses": sorted(set(stage_group["RawTicketStatus"])),
                "aging": build_aging_counts(stage_group),
            }
    stage_order = workflow_stages_for_type(ticket_type_filter)
    if not stage_order:
        stage_order = sorted(grouped, key=lambda stage: (-grouped[stage]["qty"], stage))
    for stage in grouped:
        if stage not in stage_order:
            stage_order.append(stage)
    return [
        {
            "status": stage,
            "qty": int(grouped.get(stage, {}).get("qty", 0)),
            "quoteAmount": round(float(grouped.get(stage, {}).get("quote_amount", 0)), 2),
            "quoteAmountLabel": money(float(grouped.get(stage, {}).get("quote_amount", 0))),
            "rawStatuses": grouped.get(stage, {}).get("raw_statuses", []),
            "aging": grouped.get(stage, {}).get("aging", build_aging_counts(pd.DataFrame(columns=rows.columns))),
            "color": stage_color(stage, idx),
        }
        for idx, stage in enumerate(stage_order)
    ]


def build_daily_workflow(
    normal_tickets: pd.DataFrame,
    invoice_tickets: pd.DataFrame,
    open_tickets: pd.DataFrame,
    period: str,
    ticket_type_filter: str,
    dealer: str,
) -> dict[str, Any]:
    period_obj = pd.Period(period)
    start = period_obj.to_timestamp()
    end = period_obj.to_timestamp(how="end")
    today = pd.Timestamp.now().normalize()
    if period_obj == today.to_period("M"):
        end = today

    type_rows = filter_ticket_type(normal_tickets, ticket_type_filter)
    type_invoices = filter_ticket_type(invoice_tickets, ticket_type_filter)
    type_open = filter_ticket_type(open_tickets, ticket_type_filter)
    type_rows = type_rows[type_rows["DealerBucket"].eq(dealer)].copy()
    type_invoices = type_invoices[type_invoices["DealerBucket"].eq(dealer)].copy()
    type_open = type_open[type_open["DealerBucket"].eq(dealer)].copy()

    completed_rows = type_rows[type_rows["CompletedDate"].notna()].copy()
    completed_rows = completed_rows[
        completed_rows["CompletedDate"].dt.to_period("M").astype(str).eq(period)
    ]
    created_rows = type_rows[type_rows["CreatedMonth"].eq(period)].copy()
    period_open = type_open[type_open["CreatedDate"].le(end)].copy()
    pipeline_rows = rows_created_by(type_rows, end)

    days = pd.date_range(start=start, end=end, freq="D")
    daily = []
    for day in days:
        created_day = created_rows[created_rows["CreatedDate"].dt.normalize().eq(day)]
        completed_day = completed_rows[completed_rows["CompletedDate"].dt.normalize().eq(day)]
        open_balance = type_open[type_open["CreatedDate"].le(day)]
        daily.append(
            {
                "date": day.strftime("%Y-%m-%d"),
                "label": day.strftime("%d %b"),
                "created": int(len(created_day)),
                "createdAmount": round(float(created_day["NewTicketQuoteAmount"].sum()), 2),
                "createdAmountLabel": money(float(created_day["NewTicketQuoteAmount"].sum())),
                "completed": int(len(completed_day)),
                "completedAmount": round(float(completed_day["NewTicketQuoteAmount"].sum()), 2),
                "completedAmountLabel": money(float(completed_day["NewTicketQuoteAmount"].sum())),
                "openBalance": int(len(open_balance)),
                "openAmount": round(float(open_balance["NewTicketQuoteAmount"].sum()), 2),
                "openAmountLabel": money(float(open_balance["NewTicketQuoteAmount"].sum())),
            }
        )

    return {
        "dealer": dealer,
        "ticketType": ticket_type_filter,
        "period": pd.Period(period).to_timestamp().strftime("%b %Y"),
        "daily": daily,
        "pipeline": build_status_pipeline(pipeline_rows, ticket_type_filter),
        "totals": {
            "created": int(len(created_rows)),
            "completed": int(len(completed_rows)),
            "open": int(len(period_open)),
            "createdAmountLabel": money(float(created_rows["NewTicketQuoteAmount"].sum())),
            "completedAmountLabel": money(float(completed_rows["NewTicketQuoteAmount"].sum())),
            "openAmountLabel": money(float(period_open["NewTicketQuoteAmount"].sum())),
        },
    }


def write_abnormal_workbook(abnormal: pd.DataFrame) -> None:
    export = abnormal.copy()
    preferred = [
        "AbnormalReason",
        "AbnormalMonth",
        "TicketID",
        "TicketType",
        "TicketTypeText",
        "DealerName",
        "ERPInvoiceNumber",
        "ERPInvoiceNumberPrice",
        "Billing date",
        "AmountIncludingTax",
        "CreatedOn",
        "TicketStatusText",
        "Role_1001_InvolvedPartyID",
        "Role_1001_InvolvedPartyName",
        "Role_40_InvolvedPartyName",
        "WarrantyHandlingDealerID",
        "OriginalDealerName",
        "Subject",
        "Description",
    ]
    columns = [col for col in preferred if col in export.columns]
    columns += [col for col in export.columns if col not in columns]
    export = export[columns]

    with pd.ExcelWriter(ABNORMAL_WORKBOOK, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name="AbnormalTickets")
        ws = writer.book["AbnormalTickets"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = PatternFill(fill_type="solid", fgColor="263746")
            cell.font = Font(color="FFFFFF", bold=True)
        for row_idx in range(2, ws.max_row + 1):
            reason = str(ws.cell(row=row_idx, column=1).value or "")
            if "Invoice exists but status is not Create invoice" in reason:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = ABNORMAL_FILL
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
            ws.column_dimensions[column_cells[0].column_letter].width = width


def build_dashboard_payload() -> dict[str, Any]:
    tickets = pd.read_excel(SOURCE_WORKBOOK, sheet_name="Tickets", dtype=str).fillna("")
    created = pd.to_datetime(tickets["CreatedOn"], errors="coerce", format="mixed", dayfirst=True)
    billing = pd.to_datetime(tickets["Billing date"], errors="coerce", format="mixed", dayfirst=True)
    changed = pd.to_datetime(tickets["ChangeOnDateTime"], errors="coerce", format="mixed", dayfirst=True)
    tickets["CreatedDate"] = created
    tickets["BillingDate"] = billing
    tickets["ChangeDate"] = changed
    tickets["CreatedMonth"] = created.dt.to_period("M").astype(str).where(created.notna(), "")
    tickets["CreatedYear"] = created.dt.year.astype("Int64").astype(str).where(created.notna(), "")
    tickets["BillingMonth"] = billing.dt.to_period("M").astype(str).where(billing.notna(), "")
    tickets["DealerBucket"] = tickets["DealerName"].map(dealer_bucket)
    tickets["NewTicketQuoteAmount"] = parse_amount(tickets["AmountIncludingTax"])
    tickets["InvoiceAmount"] = parse_amount(tickets["ERPInvoiceNumberPrice"])
    tickets["LabourHours"] = parse_hours(tickets["TotalLabourHours"]) if "TotalLabourHours" in tickets.columns else 0.0
    tickets["TicketTypeBucket"] = tickets.apply(ticket_type_bucket, axis=1)
    tickets["WorkerName"] = tickets.get("Role_40_InvolvedPartyName", pd.Series(dtype="object")).map(lambda value: clean(value) or "Unassigned")
    tickets["AbnormalParty"] = tickets.apply(is_abnormal_party, axis=1)
    tickets["InvoiceScope"] = tickets.apply(invoice_scope, axis=1)
    tickets["AbnormalReason"] = tickets.apply(lambda row: "; ".join(abnormal_reasons(row)), axis=1)
    tickets["AbnormalMonth"] = tickets["BillingMonth"].where(tickets["BillingMonth"].ne(""), tickets["CreatedMonth"])
    abnormal_tickets = tickets[tickets["AbnormalReason"].ne("")].copy()
    write_abnormal_workbook(abnormal_tickets)

    normal_tickets = tickets[~tickets["AbnormalParty"]].copy()
    normal_tickets["CompletedDate"] = normal_tickets.apply(completed_date, axis=1)
    created_tickets = normal_tickets[normal_tickets["CreatedMonth"].ne("")].copy()
    open_tickets = tickets[
        tickets["CreatedMonth"].ne("")
        & tickets.apply(is_open_ticket, axis=1)
    ].copy()
    missing_invoice_mask = (
        normal_tickets["TicketStatusText"].map(lambda value: clean(value).lower()).eq(CREATE_INVOICE_STATUS)
        & (
            normal_tickets["ERPInvoiceNumber"].map(clean).eq("")
            | normal_tickets["ERPInvoiceNumberPrice"].map(clean).eq("")
        )
    )
    normal_tickets["MissingInvoicePlaceholder"] = missing_invoice_mask
    valid_invoice_mask = (
        normal_tickets["BillingMonth"].ne("")
        & normal_tickets["ERPInvoiceNumber"].map(clean).ne("")
        & normal_tickets["ERPInvoiceNumberPrice"].map(clean).ne("")
        & ~normal_tickets["TicketStatusText"].map(lambda value: clean(value).lower()).isin(CANCELLED_STATUS_TEXT)
    )
    invoice_tickets = pd.concat(
        [normal_tickets[valid_invoice_mask], normal_tickets[missing_invoice_mask]],
        ignore_index=True,
    ).drop_duplicates(subset=["TicketID"], keep="first")
    invoice_tickets["InvoiceMonth"] = invoice_tickets["BillingMonth"].where(
        invoice_tickets["BillingMonth"].ne(""),
        invoice_tickets["CreatedMonth"],
    )
    invoice_tickets["InvoiceYear"] = invoice_tickets["InvoiceMonth"].str.slice(0, 4)
    invoice_tickets.loc[invoice_tickets["MissingInvoicePlaceholder"], "InvoiceAmount"] = 0.0

    all_month_periods = sorted(
        set(created_tickets["CreatedMonth"].dropna().unique())
        | set(invoice_tickets["InvoiceMonth"].dropna().unique()),
        reverse=True,
    )
    current_calendar_month = pd.Timestamp.now().to_period("M").strftime("%Y-%m")
    month_periods = [period for period in all_month_periods if period <= current_calendar_month]
    year_labels = sorted(
        set(created_tickets["CreatedYear"].dropna().unique())
        | set(invoice_tickets["InvoiceYear"].dropna().unique()),
        reverse=True,
    )
    year_labels = [year for year in year_labels if year]
    selected_month = current_calendar_month if current_calendar_month in month_periods else (month_periods[0] if month_periods else "")
    selected_label = pd.Period(selected_month).to_timestamp().strftime("%b %Y") if selected_month else ""
    other_month_periods = [period for period in month_periods if period != selected_month]
    other_month_labels = [pd.Period(m).to_timestamp().strftime("%b %Y") for m in other_month_periods]
    months = [selected_label, *year_labels, *other_month_labels] if selected_label else year_labels
    current_created = created_tickets[created_tickets["CreatedMonth"].eq(selected_month)].copy()
    current_invoiced = invoice_tickets[invoice_tickets["InvoiceMonth"].eq(selected_month)].copy()
    current_open = open_rows_at_month_end(open_tickets, selected_month) if selected_month else open_tickets.copy()

    total_qty = int(len(current_created))
    total_amount = float(current_created["NewTicketQuoteAmount"].sum())
    total_invoiced_qty = int(len(current_invoiced))
    total_invoiced_amount = float(current_invoiced["InvoiceAmount"].sum())
    total_open_qty = int(len(current_open))
    total_open_amount = float(current_open["NewTicketQuoteAmount"].sum())

    dealer_rows = build_dealer_rows(current_created, current_invoiced, current_open)

    monthly_dealer_activity = {}
    monthly_dealer_activity_by_type = {}
    monthly_invoice_mix = {}
    monthly_open_status_mix = {}
    monthly_labour = {}
    workflow_daily = {}
    for year in year_labels:
        year_created = created_tickets[created_tickets["CreatedYear"].eq(year)]
        year_invoice = invoice_tickets[invoice_tickets["InvoiceYear"].eq(year)]
        year_open = open_rows_at_year_end(open_tickets, year)
        year_labour = tickets[tickets["CreatedYear"].eq(year)].copy()
        monthly_dealer_activity_by_type[year] = {}
        for type_filter in TICKET_TYPE_FILTERS:
            monthly_dealer_activity_by_type[year][type_filter] = build_dealer_rows(
                filter_ticket_type(year_created, type_filter),
                filter_ticket_type(year_invoice, type_filter),
                filter_ticket_type(year_open, type_filter),
                type_filter,
            )
        monthly_dealer_activity[year] = monthly_dealer_activity_by_type[year]["All Ticket Types"]
        monthly_invoice_mix[year] = build_invoice_mix(year_invoice)
        monthly_open_status_mix[year] = build_open_status_mix(year_open)
        monthly_labour[year] = build_labour_summary(year_labour)
    for period in month_periods:
        label = pd.Period(period).to_timestamp().strftime("%b %Y")
        period_open = open_rows_at_month_end(open_tickets, period)
        period_created = created_tickets[created_tickets["CreatedMonth"].eq(period)]
        period_invoice = invoice_tickets[invoice_tickets["InvoiceMonth"].eq(period)]
        period_labour = tickets[tickets["CreatedMonth"].eq(period)].copy()
        monthly_dealer_activity_by_type[label] = {}
        for type_filter in TICKET_TYPE_FILTERS:
            monthly_dealer_activity_by_type[label][type_filter] = build_dealer_rows(
                filter_ticket_type(period_created, type_filter),
                filter_ticket_type(period_invoice, type_filter),
                filter_ticket_type(period_open, type_filter),
                type_filter,
            )
        monthly_dealer_activity[label] = monthly_dealer_activity_by_type[label]["All Ticket Types"]
        monthly_invoice_mix[label] = build_invoice_mix(invoice_tickets[invoice_tickets["InvoiceMonth"].eq(period)])
        monthly_open_status_mix[label] = build_open_status_mix(period_open)
        monthly_labour[label] = build_labour_summary(period_labour)
        workflow_daily[label] = {}
        for type_filter in TICKET_TYPE_FILTERS:
            if type_filter == "All Ticket Types":
                continue
            workflow_daily[label][type_filter] = {
                dealer: build_daily_workflow(
                    normal_tickets,
                    invoice_tickets,
                    open_tickets,
                    period,
                    type_filter,
                    dealer,
                )
                for dealer in DEALER_ORDER
                if dealer != "Other"
            }

    trend_rows = []
    for period in month_periods[:12]:
        created_rows = created_tickets[created_tickets["CreatedMonth"].eq(period)]
        invoice_rows = invoice_tickets[invoice_tickets["InvoiceMonth"].eq(period)]
        open_rows = open_rows_at_month_end(open_tickets, period)
        trend_rows.append(
            {
                "month": pd.Period(period).to_timestamp().strftime("%b"),
                "monthFull": pd.Period(period).to_timestamp().strftime("%b %Y"),
                "newTickets": int(len(created_rows)),
                "newAmount": round(float(created_rows["NewTicketQuoteAmount"].sum()), 2),
                "newQuoteAmount": round(float(created_rows["NewTicketQuoteAmount"].sum()), 2),
                "invoicedTickets": int(len(invoice_rows)),
                "invoicedAmount": round(float(invoice_rows["InvoiceAmount"].sum()), 2),
                "openTickets": int(len(open_rows)),
                "openQuoteAmount": round(float(open_rows["NewTicketQuoteAmount"].sum()), 2),
            }
        )
    trend_rows.reverse()

    generated_at = pd.Timestamp.now().strftime("%d %b %Y, %I:%M %p")

    return {
        "meta": {
            "lastUpdated": generated_at,
            "months": months,
            "yards": ["All Dealers", *DEALER_ORDER],
            "invoiceScopes": ["All Invoices", "Internal", "External"],
            "ticketTypes": TICKET_TYPE_FILTERS,
            "currentMonth": selected_label,
            "abnormalExportFile": "../abnormal_tickets.xlsx",
            "abnormalTickets": int(len(abnormal_tickets)),
        },
        "pages": {
            "overview": {
                "title": "Service Centre Overview",
                "subtitle": "C4C new ticket creation by dealer, based on CreatedOn.",
                "kpis": [
                    {
                        "title": "New Tickets Created",
                        "value": f"{total_qty:,}",
                        "detail": f"{selected_label} total across five target dealers plus Other",
                        "icon": "NT",
                        "tone": "blue",
                        "badge": selected_label,
                        "badgeTone": "up",
                    },
                    {
                        "title": "New Ticket Quote Amount",
                        "value": money(total_amount),
                        "detail": "AmountIncludingTax from exported C4C ticket result",
                        "icon": "AM",
                        "tone": "teal",
                        "badge": "CreatedOn",
                        "badgeTone": "up",
                    },
                    {
                        "title": "Target Dealers",
                        "value": "5",
                        "detail": "Perth, Traralgon, Launceston, Geelong and Frankston",
                        "icon": "DL",
                        "tone": "green",
                        "badge": "+ Other",
                        "badgeTone": "warn",
                    },
                    {
                        "title": "Other Tickets",
                        "value": f"{next(r['newTickets'] for r in dealer_rows if r['yard'] == 'Other'):,}",
                        "detail": "All dealers outside the five target dealer buckets",
                        "icon": "OT",
                        "tone": "orange",
                        "badge": next(r["newAmountLabel"] for r in dealer_rows if r["yard"] == "Other"),
                        "badgeTone": "warn",
                    },
                ],
                "yardActivity": dealer_rows,
                "monthlyDealerActivity": monthly_dealer_activity,
                "monthlyDealerActivityByType": monthly_dealer_activity_by_type,
                "invoiceMix": build_invoice_mix(current_invoiced),
                "monthlyInvoiceMix": monthly_invoice_mix,
                "openStatusMix": build_open_status_mix(current_open),
                "monthlyOpenStatusMix": monthly_open_status_mix,
                "monthlyLabour": monthly_labour,
                "ticketDetails": build_ticket_details(created_tickets),
                "workflowDaily": workflow_daily,
                "yardSummary": [
                    {
                        "yard": row["yard"],
                        "color": row["color"],
                        "newTickets": row["newTickets"],
                        "newAmount": row["newAmount"],
                        "newAmountLabel": row["newAmountLabel"],
                        "status": "Target" if row["yard"] != "Other" else "Other",
                        "statusTone": "good" if row["yard"] != "Other" else "watch",
                    }
                    for row in dealer_rows
                ],
                "trend": trend_rows,
            },
            "activity": {
                "title": "Ticket Activity",
                "subtitle": "Monthly new ticket creation by CreatedOn.",
                "kpis": [
                    {
                        "title": "Current Month Qty",
                        "value": f"{total_qty:,}",
                        "detail": f"New tickets created in {selected_label}",
                        "icon": "NQ",
                        "tone": "blue",
                        "badge": selected_label,
                        "badgeTone": "up",
                    },
                    {
                        "title": "Current Month Quote Amount",
                        "value": money(total_amount),
                        "detail": "AmountIncludingTax total for CreatedOn month",
                        "icon": "NA",
                        "tone": "teal",
                        "badge": "C4C",
                        "badgeTone": "up",
                    },
                ],
                "trend": trend_rows,
                "breakdown": [
                    {
                        "title": row["yard"],
                        "amount": row["newAmountLabel"],
                        "count": f"{row['newTickets']:,} tickets",
                        "note": "Current CreatedOn month",
                    }
                    for row in dealer_rows
                ],
                "insights": [],
            },
            "backlog": {"title": "Open Backlog", "subtitle": "Pending next KPI build.", "kpis": [], "rows": [], "issues": []},
            "hours": {"title": "Hours & Payroll", "subtitle": "Pending HR data source.", "kpis": [], "rows": [], "balance": []},
        },
    }


def main() -> None:
    payload = build_dashboard_payload()
    OUTPUT_JS.write_text(
        "window.serviceCentreData = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Dashboard data written: {OUTPUT_JS}")


if __name__ == "__main__":
    main()
