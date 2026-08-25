from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
SOURCE_WORKBOOK = BASE_DIR / "c4c_ticket_table_z007_z010_checked_hana_final.xlsx"
SAP_LABOUR_REPORT = Path(
    r"C:\Users\Leo.Li\Downloads\SAPAnalyticsReport(Z3F22153042C7B7B40972C2) (1).xlsx"
)
SUMMARY_JSON = Path("outputs") / "sap_labour_patch_summary.json"


def clean_ticket_id(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def clean_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    return "" if text.lower() == "nan" else text


def load_sap_labour() -> pd.DataFrame:
    raw = pd.read_excel(SAP_LABOUR_REPORT, sheet_name=0, header=None, dtype=str).fillna("")
    labour = raw.iloc[10:, [0, 2, 6]].copy()
    labour.columns = ["TicketID", "Worker", "TotalLabourHours"]
    labour["TicketID"] = labour["TicketID"].map(clean_ticket_id)
    labour = labour[pd.to_numeric(labour["TicketID"], errors="coerce").notna()].copy()
    labour["Worker"] = labour["Worker"].map(clean_text)
    labour["TotalLabourHours"] = pd.to_numeric(
        labour["TotalLabourHours"].astype(str).str.strip(),
        errors="coerce",
    ).fillna(0.0)

    grouped = (
        labour.groupby("TicketID", as_index=False)
        .agg(
            TotalLabourHours=("TotalLabourHours", "sum"),
            Worker=("Worker", lambda values: "; ".join(dict.fromkeys(v for v in values if v))),
        )
    )
    return grouped


def header_map(ws) -> dict[str, int]:
    result: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            result[str(cell.value).strip()] = cell.column
    return result


def ensure_column(ws, headers: dict[str, int], column_name: str, after_column: str | None = None) -> int:
    if column_name in headers:
        return headers[column_name]

    insert_at = ws.max_column + 1
    if after_column and after_column in headers:
        insert_at = headers[after_column] + 1
        ws.insert_cols(insert_at)

    ws.cell(1, insert_at).value = column_name
    return insert_at


def main() -> None:
    labour = load_sap_labour()
    labour_by_ticket = labour.set_index("TicketID").to_dict("index")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    backup_path = SOURCE_WORKBOOK.with_name(
        f"{SOURCE_WORKBOOK.stem}_backup_before_sap_labour_{timestamp}{SOURCE_WORKBOOK.suffix}"
    )
    shutil.copy2(SOURCE_WORKBOOK, backup_path)

    wb = load_workbook(SOURCE_WORKBOOK)
    ws = wb["Tickets"]
    headers = header_map(ws)
    ensure_column(ws, headers, "TotalLabourHours", after_column="Z1Z8TimeConsumed")
    headers = header_map(ws)
    required = ["TicketID", "Role_40_InvolvedPartyName", "TotalLabourHours"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f"Missing required columns in Tickets sheet: {missing}")

    ticket_col = headers["TicketID"]
    worker_col = headers["Role_40_InvolvedPartyName"]
    hours_col = headers["TotalLabourHours"]

    matched = 0
    changed_hours = 0
    changed_workers = 0
    old_hours_sum = 0.0
    new_hours_sum = 0.0
    unmatched_c4c = []

    for row in range(2, ws.max_row + 1):
        ticket_id = clean_ticket_id(ws.cell(row, ticket_col).value)
        old_worker = clean_text(ws.cell(row, worker_col).value)
        old_hours = pd.to_numeric(ws.cell(row, hours_col).value, errors="coerce")
        old_hours = 0.0 if pd.isna(old_hours) else float(old_hours)
        old_hours_sum += old_hours

        patch = labour_by_ticket.get(ticket_id)
        if patch is None:
            unmatched_c4c.append(ticket_id)
            new_hours_sum += old_hours
            continue

        matched += 1
        new_worker = clean_text(patch["Worker"])
        new_hours = float(patch["TotalLabourHours"])
        if abs(old_hours - new_hours) > 1e-9:
            changed_hours += 1
        if old_worker != new_worker:
            changed_workers += 1

        ws.cell(row, worker_col).value = new_worker
        ws.cell(row, hours_col).value = round(new_hours, 4)
        ws.cell(row, hours_col).number_format = "0.0000"
        new_hours_sum += new_hours

    wb.save(SOURCE_WORKBOOK)

    c4c_ticket_ids = {
        clean_ticket_id(ws.cell(row, ticket_col).value)
        for row in range(2, ws.max_row + 1)
    }
    sap_not_in_c4c = sorted(set(labour_by_ticket) - c4c_ticket_ids)
    summary = {
        "sourceWorkbook": str(SOURCE_WORKBOOK),
        "sapLabourReport": str(SAP_LABOUR_REPORT),
        "backupWorkbook": str(backup_path),
        "sapRowsAfterGrouping": int(len(labour_by_ticket)),
        "matchedTickets": matched,
        "unmatchedC4CTickets": len(unmatched_c4c),
        "sapTicketsNotInC4C": len(sap_not_in_c4c),
        "changedHours": changed_hours,
        "changedWorkers": changed_workers,
        "oldHoursSum": round(old_hours_sum, 4),
        "newHoursSum": round(new_hours_sum, 4),
        "sapMatchedHoursSum": round(float(labour[labour["TicketID"].isin(c4c_ticket_ids)]["TotalLabourHours"].sum()), 4),
        "sapTicketsNotInC4CSample": sap_not_in_c4c[:20],
    }
    SUMMARY_JSON.parent.mkdir(parents=True, exist_ok=True)
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
