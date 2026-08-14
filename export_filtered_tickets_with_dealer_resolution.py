from __future__ import annotations

import logging
import os
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, List, Tuple
from urllib.parse import quote

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from requests.auth import HTTPBasicAuth
from urllib3.util.retry import Retry

# =================== C4C API configuration ===================
BASE_URL = os.getenv(
    "C4C_BASE_URL",
    "https://longcui-automobile-cpi-tyrbc1k7.it-cpi010-rt.cpi.cn40.apps.platform.sapcloud.cn",
)
PATH = os.getenv("C4C_API_PATH", "/http/PC4C/Ticket/queryOdataBatch")

USERNAME = os.getenv("C4C_USERNAME", "")
PASSWORD = os.getenv("C4C_PASSWORD", "")

ROLE_CODES = [x.strip() for x in os.getenv("C4C_ROLE_CODES", "1001,40,43").split(",") if x.strip()]
API_TOP = int(os.getenv("C4C_API_TOP", "1000"))
API_SKIP_START = int(os.getenv("C4C_API_SKIP_START", "0"))
TIMEOUT = int(os.getenv("C4C_TIMEOUT", "60"))
VERIFY_SSL = os.getenv("C4C_VERIFY_SSL", "true").lower() in {"1", "true", "yes", "y"}
MAX_WORKERS = int(os.getenv("C4C_MAX_WORKERS", "12"))
ROLE_WORKERS = min(max(1, len(ROLE_CODES)), 3)

OUTPUT_FILE = os.getenv(
    "OUTPUT_FILE",
    "c4c_ticket_table_z007_z010_with_invoice_layout_checked.xlsx",
)
SOURCE_TICKET_FILE = os.getenv("SOURCE_TICKET_FILE", "")
SHEET_NAME = "Tickets"
TARGET_TICKET_TYPES = {"Z007", "Z010"}
EXCLUDED_WARRANTY_DEALER_IDS = {"3110", "3151"}
# =============================================================

ROLE_VARYING_FIELDS = {
    "InvolvedPartyBusinessPartnerID",
    "InvolvedPartyID",
    "InvolvedPartyName",
    "InvolvedPartyRoleID",
    "requested_skip",
}

REQUEST_META_FIELDS = {
    "requested_role_code",
    "requested_role_name",
    "requested_skip",
}

MANUAL_WARRANTY_DEALER_MAPPING = {
    "3113": "Regent RV Pty Ltd",
    "311301": "Regent RV Pty Ltd",
    "3121": "Regent RV - Perth",
    "312201": "Regent RV - Perth",
    "3123": "Regent RV - Traralgon",
    "312401": "Regent RV - Traralgon",
    "3126": "Snowy River Launceston",
    "312701": "Snowy River Launceston",
    "3128": "Snowy River Geelong",
    "312901": "Snowy River Geelong",
    "3141": "Regent RV - Frankston",
    "314201": "Regent RV - Frankston",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("c4c_ticket_filtered_export")
_thread_local = threading.local()


def get_thread_session() -> requests.Session:
    session = getattr(_thread_local, "session", None)
    if session is None:
        session = requests.Session()
        session.trust_env = False
        retry = Retry(
            total=5,
            backoff_factor=0.6,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"],
            raise_on_status=False,
        )
        adapter = HTTPAdapter(
            max_retries=retry,
            pool_connections=50,
            pool_maxsize=50,
        )
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        _thread_local.session = session
    return session


def close_thread_session() -> None:
    session = getattr(_thread_local, "session", None)
    if session is not None:
        try:
            session.close()
        finally:
            _thread_local.session = None


def normalize_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, (dict, list)):
        return str(value)
    return value


def build_url(role_code: str, top: int, skip: int) -> str:
    filter_text = f"(CCSRQ_DPY_ROLE_CD eq '{role_code}')"
    filter_value = quote(filter_text, safe="()'")
    return (
        BASE_URL.rstrip("/")
        + PATH
        + f"?$top={top}&$skip={skip}&$filter={filter_value}"
    )


def fetch_role_page(role_code: str, top: int, skip: int) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    response = get_thread_session().get(
        build_url(role_code, top, skip),
        auth=HTTPBasicAuth(USERNAME, PASSWORD),
        headers={"Accept": "application/json"},
        timeout=TIMEOUT,
        verify=VERIFY_SSL,
    )

    if response.status_code != 200:
        raise RuntimeError(
            f"API failed: role={role_code}, skip={skip}, "
            f"HTTP={response.status_code}, body={response.text[:500]}"
        )

    payload = response.json()
    rows = list(payload.get("data", []))
    meta = {
        "count": payload.get("count"),
        "totalCount": payload.get("totalCount"),
    }
    return rows, meta


def fetch_role_page_task(role_code: str, skip: int):
    rows, meta = fetch_role_page(role_code, API_TOP, skip)
    return skip, rows, meta


def fetch_all_rows_for_role(role_code: str) -> List[Dict[str, Any]]:
    started = time.time()
    all_rows: List[Dict[str, Any]] = []

    first_rows, meta = fetch_role_page(role_code, API_TOP, API_SKIP_START)
    all_rows.extend(first_rows)
    logger.info("role=%s first page rows=%s", role_code, len(first_rows))

    if not first_rows:
        return []

    total_raw = meta.get("totalCount") or meta.get("count")
    try:
        total = int(total_raw)
    except (TypeError, ValueError):
        total = 0

    if total > 0:
        skips = list(range(API_SKIP_START + API_TOP, total, API_TOP))
        page_map: Dict[int, List[Dict[str, Any]]] = {}
        workers = min(MAX_WORKERS, max(1, len(skips)))

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(fetch_role_page_task, role_code, skip): skip
                for skip in skips
            }
            for future in as_completed(futures):
                skip, rows, _ = future.result()
                page_map[skip] = rows
                logger.info("role=%s skip=%s rows=%s", role_code, skip, len(rows))

        for skip in sorted(page_map):
            all_rows.extend(page_map[skip])
    else:
        skip = API_SKIP_START + API_TOP
        while True:
            rows, _ = fetch_role_page(role_code, API_TOP, skip)
            logger.info("role=%s skip=%s rows=%s", role_code, skip, len(rows))
            if not rows:
                break
            all_rows.extend(rows)
            if len(rows) < API_TOP:
                break
            skip += API_TOP

    logger.info(
        "role=%s complete, rows=%s, elapsed=%.1fs",
        role_code,
        len(all_rows),
        time.time() - started,
    )
    return all_rows


def merge_rows_to_ticket_table(rows_by_role: Dict[str, List[Dict[str, Any]]]) -> pd.DataFrame:
    tickets: Dict[str, Dict[str, Any]] = {}

    for role_code, rows in rows_by_role.items():
        for row in rows:
            ticket_id = str(row.get("TicketID") or "").strip()
            if not ticket_id:
                continue

            record = tickets.setdefault(ticket_id, {"TicketID": ticket_id})

            for key, value in row.items():
                if key in REQUEST_META_FIELDS or key == "TicketID":
                    continue

                clean_value = normalize_value(value)
                if key in ROLE_VARYING_FIELDS:
                    if key != "requested_skip":
                        record[f"Role_{role_code}_{key}"] = clean_value
                elif key not in record or record.get(key) in (None, ""):
                    record[key] = clean_value

    if not tickets:
        return pd.DataFrame()

    df = pd.DataFrame(tickets.values())

    preferred_columns = [
        "TicketID",
        "DealerID",
        "DealerName",
        "WarrantyHandlingDealerID",
        "TicketType",
        "TicketTypeText",
        "CreatedOn",
        "TicketStatus",
        "TicketStatusText",
        "AmountIncludingTax",
        "ERPInvoiceNumber",
        "Description",
        "Subject",
    ]
    ordered = [col for col in preferred_columns if col in df.columns]
    ordered += sorted(col for col in df.columns if col not in ordered)
    df = df[ordered]

    if "CreatedOn" in df.columns:
        parsed = pd.to_datetime(df["CreatedOn"], errors="coerce")
        formatted = parsed.dt.strftime("%Y-%m-%d %H:%M:%S")
        df["CreatedOn"] = formatted.where(parsed.notna(), df["CreatedOn"])

    return df.sort_values("TicketID", kind="stable").reset_index(drop=True)


def load_source_ticket_table() -> pd.DataFrame:
    if SOURCE_TICKET_FILE and os.path.exists(SOURCE_TICKET_FILE):
        logger.info("Loading source ticket table from local file: %s", SOURCE_TICKET_FILE)
        return pd.read_excel(SOURCE_TICKET_FILE, dtype=str)
    return pd.DataFrame()


def build_existing_dealer_map(df: pd.DataFrame) -> Dict[str, str]:
    if "DealerID" not in df.columns or "DealerName" not in df.columns:
        return {}

    source = df[["DealerID", "DealerName"]].copy()
    source["DealerID"] = source["DealerID"].fillna("").astype(str).str.strip()
    source["DealerName"] = source["DealerName"].fillna("").astype(str).str.strip()
    source = source[
        source["DealerID"].ne("")
        & source["DealerName"].ne("")
        & ~source["DealerName"].str.contains("Not Assigned", case=False, na=False)
    ]
    if source.empty:
        return {}
    return source.drop_duplicates().groupby("DealerID")["DealerName"].first().to_dict()


def apply_filters_and_resolution(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if df.empty:
        empty = pd.DataFrame()
        return empty, empty, empty, empty

    filtered = df.copy()
    filtered["TicketType"] = filtered.get("TicketType", "").fillna("").astype(str).str.strip().str.upper()
    filtered = filtered[filtered["TicketType"].isin(TARGET_TICKET_TYPES)].copy()

    existing_map = build_existing_dealer_map(df)
    dealer_map = {**existing_map, **MANUAL_WARRANTY_DEALER_MAPPING}

    filtered["DealerID"] = filtered.get("DealerID", "").fillna("").astype(str).str.strip()
    filtered["DealerName"] = filtered.get("DealerName", "").fillna("").astype(str).str.strip()
    filtered["OriginalDealerName"] = filtered["DealerName"]
    filtered["WarrantyHandlingDealerID"] = (
        filtered.get("WarrantyHandlingDealerID", "").fillna("").astype(str).str.strip()
    )
    filtered = filtered[~filtered["WarrantyHandlingDealerID"].isin(EXCLUDED_WARRANTY_DEALER_IDS)].copy()
    filtered["MappedDealerName"] = filtered["WarrantyHandlingDealerID"].map(dealer_map).fillna("")

    resolution_status: List[str] = []
    resolved_dealer_name: List[str] = []

    for _, row in filtered.iterrows():
        dealer_name = row["DealerName"]
        warranty_dealer_id = row["WarrantyHandlingDealerID"]
        mapped_name = row["MappedDealerName"]

        if warranty_dealer_id and mapped_name:
            resolution_status.append("resolved from warranty handling dealer")
            resolved_dealer_name.append(mapped_name)
        elif dealer_name and "not assigned" not in dealer_name.lower():
            resolution_status.append("original dealer present")
            resolved_dealer_name.append(dealer_name)
        elif not warranty_dealer_id:
            resolution_status.append("warranty handling dealer empty")
            resolved_dealer_name.append("")
        else:
            resolution_status.append("warranty handling dealer unmapped")
            resolved_dealer_name.append("")

    filtered["ResolvedDealerName"] = resolved_dealer_name
    filtered["FinalDealerName"] = filtered["ResolvedDealerName"].where(
        filtered["ResolvedDealerName"].ne(""),
        filtered["OriginalDealerName"],
    )
    filtered["DealerName"] = filtered["FinalDealerName"]
    if "ERPInvoiceNumberPrice" in filtered.columns:
        filtered["ERPInvoiceNumberPrice"] = filtered["ERPInvoiceNumberPrice"].fillna("").astype(str)
    else:
        filtered["ERPInvoiceNumberPrice"] = ""
    if "Billing date" in filtered.columns:
        filtered["Billing date"] = filtered["Billing date"].fillna("").astype(str)
    else:
        filtered["Billing date"] = ""
    filtered["DealerResolutionStatus"] = resolution_status

    unresolved = filtered[
        filtered["DealerResolutionStatus"].isin(
            ["warranty handling dealer empty", "warranty handling dealer unmapped"]
        )
    ].copy()
    main_tickets = filtered[
        ~filtered["DealerResolutionStatus"].isin(
            ["warranty handling dealer empty", "warranty handling dealer unmapped"]
        )
    ].copy()

    summary = (
        filtered.groupby(
            ["TicketType", "DealerResolutionStatus", "WarrantyHandlingDealerID", "ResolvedDealerName"],
            dropna=False,
        )
        .size()
        .reset_index(name="TicketCount")
        .sort_values(
            ["TicketCount", "TicketType", "WarrantyHandlingDealerID"],
            ascending=[False, True, True],
        )
    )

    mapping_sheet = pd.DataFrame(
        sorted(dealer_map.items()),
        columns=["WarrantyHandlingDealerID", "MappedDealerName"],
    )

    drop_export_columns = ["ResolvedDealerName", "FinalDealerName", "MappedDealerName"]
    filtered = filtered.drop(columns=[c for c in drop_export_columns if c in filtered.columns])
    main_tickets = main_tickets.drop(columns=[c for c in drop_export_columns if c in main_tickets.columns])
    unresolved = unresolved.drop(columns=[c for c in drop_export_columns if c in unresolved.columns])

    preferred_columns = [
        "TicketID",
        "TicketType",
        "TicketTypeText",
        "DealerID",
        "DealerName",
        "ERPInvoiceNumber",
        "ERPInvoiceNumberPrice",
        "Billing date",
        "AmountIncludingTax",
        "WarrantyHandlingDealerID",
        "CreatedOn",
        "TicketStatus",
        "TicketStatusText",
        "ERPFreeOrder",
        "Role_43_InvolvedPartyName",
        "TicketName",
        "SerialID",
        "ChassisNumber",
        "DealerResolutionStatus",
        "OriginalDealerName",
    ]
    existing_columns = [col for col in preferred_columns if col in filtered.columns]
    ordered_columns = existing_columns + [col for col in filtered.columns if col not in existing_columns]
    filtered = filtered[ordered_columns]
    main_tickets = main_tickets[[col for col in ordered_columns if col in main_tickets.columns]]
    unresolved = unresolved[[col for col in ordered_columns if col in unresolved.columns]]

    return (
        main_tickets.reset_index(drop=True),
        unresolved.reset_index(drop=True),
        summary.reset_index(drop=True),
        mapping_sheet,
    )


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 24

    for col_idx, column_name in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        header = str(column_name[0].value or "")
        values = [header]
        for cell in ws[get_column_letter(col_idx)][1:1001]:
            values.append("" if cell.value is None else str(cell.value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        width = min(max(max(len(v) for v in values) + 2, 12), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def export_excel(
    tickets_df: pd.DataFrame,
    not_assigned_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    mapping_df: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        tickets_df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        not_assigned_df.to_excel(writer, index=False, sheet_name="NotAssigned")
        summary_df.to_excel(writer, index=False, sheet_name="DealerMappingResult")
        mapping_df.to_excel(writer, index=False, sheet_name="DealerMappingUsed")

        for sheet_name in writer.sheets:
            format_sheet(writer.sheets[sheet_name])

    logger.info("Excel created: %s (%s ticket rows)", OUTPUT_FILE, len(tickets_df))


def main() -> None:
    started = time.time()
    rows_by_role: Dict[str, List[Dict[str, Any]]] = {}

    try:
        ticket_df = load_source_ticket_table()
        if ticket_df.empty:
            if not USERNAME or not PASSWORD:
                raise SystemExit(
                    "Please set C4C_USERNAME and C4C_PASSWORD environment variables before running."
                )
            with ThreadPoolExecutor(max_workers=ROLE_WORKERS) as executor:
                futures = {
                    executor.submit(fetch_all_rows_for_role, role_code): role_code
                    for role_code in ROLE_CODES
                }
                for future in as_completed(futures):
                    role_code = futures[future]
                    rows_by_role[role_code] = future.result()

            ticket_df = merge_rows_to_ticket_table(rows_by_role)

        tickets_df, not_assigned_df, summary_df, mapping_df = apply_filters_and_resolution(ticket_df)
        export_excel(tickets_df, not_assigned_df, summary_df, mapping_df)
        logger.info(
            "Done. Ticket rows=%s not_assigned=%s elapsed=%.1fs",
            len(tickets_df),
            len(not_assigned_df),
            time.time() - started,
        )
    finally:
        close_thread_session()


if __name__ == "__main__":
    main()
